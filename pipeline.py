import os
import shutil
import subprocess
import sys

from openpyxl import Workbook, load_workbook
from faster_whisper import WhisperModel
from sarvamai import SarvamAI
from saaras_stt import transcribe as saaras_transcribe
from sarvamai.play import save



# ============================================================
#                       CONSTANTS
# ============================================================

def get_whisper_model():

    print("\nWhisper Model")
    print("1. Small")
    print("2. Medium")
    

    choice = ask_choice(
        "\nEnter choice : ",
        {"1", "2"}
    )

    models = {
        "1": "small",
        "2": "medium",

    }

    return models[choice]

def get_stt_engine():

    print("\nSpeech To Text Engine")
    print("1. Whisper")
    print("2. Saaras v4")

    choice = ask_choice(
        "\nEnter choice : ",
        {"1", "2"}
    )

    if choice == "1":
        return "whisper"

    return "saaras"

MASTER_EXCEL = "Master_Transcript.xlsx"

MASTER_HEADERS = [
    "File Name",
    "Source Transcript"
]

LANGUAGE_HEADERS = [
    "File Name",
    "Source Transcript",
    "Translation"
]


AUDIO_EXTENSIONS = (
    ".wav",
    ".mp3",
    ".flac",
    ".aac",
    ".m4a",
    ".ogg"
)

LANGUAGES = {
    
    "0": {
        "name": "English",
        "code": "en-IN",
        "prefix": "en"
    },
    "1": {
        "name": "Hindi",
        "code": "hi-IN",
        "prefix": "hi"
    },
    "2": {
        "name": "Telugu",
        "code": "te-IN",
        "prefix": "te"
    },
    "3": {
        "name": "Marathi",
        "code": "mr-IN",
        "prefix": "mr"
    },
    "4": {
        "name": "Punjabi",
        "code": "pa-IN",
        "prefix": "pa"
    },
    "5": {
        "name": "Tamil",
        "code": "ta-IN",
        "prefix": "ta"
    },
    "6": {
        "name": "Bengali",
        "code": "bn-IN",
        "prefix": "bn"
    },
    "7": {
        "name": "Kannada",
        "code": "kn-IN",
        "prefix": "kn"
    },
    "8": {
        "name": "Gujarati",
        "code": "gu-IN",
        "prefix": "gu"
    }
 }

TRANSLATION_MODES = {
    "1": {
        "name": "Formal",
        "value": "formal"
    },
    "2": {
        "name": "Classic Colloquial",
        "value": "classic-colloquial"
    },
    "3": {
        "name": "Modern Colloquial",
        "value": "modern-colloquial"
    },
    "4": {
        "name": "Code Mixed",
        "value": "code-mixed"
    }
}

# ============================================================
#                     INPUT HELPERS
# ============================================================

def ask_non_empty(message):

    while True:

        value = input(message).strip()

        if value:
            return value

        print("Cannot be empty.")


def ask_choice(message, choices):

    while True:

        value = input(message).strip()

        if value in choices:
            return value

        print("Invalid choice.")
        
        
def get_pace():
    while True:
        try:
            value = input("Pace (Default 1.0): ").strip()
            if value == "":
                return 1.0
            value = float(value)
            if 0.5 <= value <= 2.0:
                return value
        except ValueError:
            pass
        print("Enter a value between 0.5 and 2.0.")


# ============================================================
#                     USER INPUT
# ============================================================

def get_api_key():

    return ask_non_empty("\nSarvam API Key : ")


def get_input_type():

    print("\nInput Type")
    print("1. Audio")
    print("2. Transcript")

    choice = ask_choice(
        "\nEnter choice : ",
        {"1", "2"}
    )

    return "audio" if choice == "1" else "transcript"


def get_project_folder():

    while True:

        folder = input(
            "\nProject Folder : "
        ).strip().strip('"')

        if os.path.isdir(folder):
            return os.path.abspath(folder)

        print("Folder not found.")
        
def get_transcript_excel():

    while True:

        excel = input(
            "\nTranscript Excel Path : "
        ).strip().strip('"')

        if not os.path.isfile(excel):
            print("File not found.")
            continue

        if not excel.lower().endswith(".xlsx"):
            print("Please select an .xlsx file.")
            continue

        return os.path.abspath(excel)


def get_translation_mode():

    print("\nTranslation Style\n")

    for key, value in TRANSLATION_MODES.items():
        print(f"{key}. {value['name']}")

    choice = ask_choice(
        "\nEnter choice : ",
        TRANSLATION_MODES.keys()
    )

    return TRANSLATION_MODES[choice]["value"]

def get_transcript_language():

    print("\nTranscript Language\n")

    for key, value in LANGUAGES.items():
        print(f"{key}. {value['name']}")

    choice = ask_choice(
        "\nEnter choice : ",
        LANGUAGES.keys()
    )

    return LANGUAGES[choice]["code"]


def get_target_languages():

    print("\nTarget Languages\n")

    for key, value in LANGUAGES.items():

        print(f"{key}. {value['name']}")

    while True:

        raw = input(
            "\nEnter comma separated values : "
        ).replace(" ", "")

        ids = raw.split(",")

        if len(ids) != len(set(ids)):
            print("Duplicate languages selected.")
            continue

        if all(i in LANGUAGES for i in ids):
            return ids

        print("Invalid selection.")

def get_audio_format():

    print("\nOutput Audio Format")

    print("1. OGG (Recommended for Unity)")
    print("2. MP3")
    print("3. WAV")

    choice = ask_choice(
        "\nEnter choice : ",
        {"1", "2", "3"}
    )

    formats = {
        "1": "ogg",
        "2": "mp3",
        "3": "wav"
    }

    return formats[choice]

def configure_languages(language_ids, transcript_excel=None, input_type="audio"):

    jobs = []

    for language_id in language_ids:

        lang = LANGUAGES[language_id]

        print(f"\n========== {lang['name']} ==========")

        mode = ask_choice(
            "Mode (T / A / TA) : ",
            {"T", "A", "TA"}
        ).upper()

        if input_type == "transcript":

            source_wb = load_workbook(transcript_excel)
            source_ws = source_wb.active
            source_headers = validate_file_name_header(source_ws)
            source_wb.close()

            text_column = select_text_column(source_headers)

        else:

            text_column = "Source Transcript"

        speaker = None

        if mode in ("A", "TA"):

            speaker = ask_non_empty(
                "Speaker : "
            )
            pace = get_pace()
            
        else: 
            pace = 1.0

        jobs.append ({

            "name": lang["name"],
            "code": lang["code"],
            "prefix": lang["prefix"],
            "mode": mode,
            "speaker": speaker,
            "pace": pace,
            "text_column": text_column

        })

    return jobs


# ============================================================
#                  PROJECT HELPERS
# ============================================================

def master_excel(project):

    return os.path.join(
        project,
        MASTER_EXCEL
    )


def language_folder(project, job):

    return os.path.join(
        project,
        job["name"]
    )


def language_excel(project, job):

    return os.path.join(
        language_folder(project, job),
        f"{job['name']}.xlsx"
    )


def audio_folder(project, job):

    return os.path.join(
        language_folder(project, job),
        "Audio"
    )


def preview_folder(project, job):

    return os.path.join(
        audio_folder(project, job),
        "Preview"
    )


def project_exists(project):

    return os.path.exists(
        master_excel(project)
    )


def get_project_mode(project):

    if not project_exists(project):
        return "fresh"

    print("\nExisting Project Found")
    print("1. Resume")
    print("2. Start Fresh")

    choice = ask_choice(
        "\nEnter choice : ",
        {"1", "2"}
    )

    if choice == "1":
        return "resume"

    if os.path.exists(master_excel(project)):
        os.remove(master_excel(project))

    for lang in LANGUAGES.values():

        folder = os.path.join(
            project,
            lang["name"]
        )

        if os.path.exists(folder):
            shutil.rmtree(folder)

    return "fresh"


def create_output_structure(project, jobs):

    for job in jobs:

        folder = language_folder(project, job)

        os.makedirs(
            folder,
            exist_ok=True
        )

        if job["mode"] in ("A", "TA"):

            os.makedirs(
                audio_folder(project, job),
                exist_ok=True
            )

            os.makedirs(
                preview_folder(project, job),
                exist_ok=True
            )

        job["folder"] = folder
        job["excel"] = language_excel(project, job)
        job["audio_folder"] = audio_folder(project, job)
        job["preview_folder"] = preview_folder(project, job)

# ============================================================
#                     VALIDATION
# ============================================================

def validate_audio_project(project):

    files = [

        f for f in os.listdir(project)

        if f.lower().endswith(AUDIO_EXTENSIONS)

    ]

    if not files:

        raise Exception(
            "No audio files found in project folder."
        )

def validate_headers(ws, required):

    headers = get_worksheet_headers(ws)

    for item in required:

        if item not in headers:

            raise Exception(
                f"Missing header : {item}"
            )

    return headers


def get_worksheet_headers(ws):

    headers = {}

    for col in range(1, ws.max_column + 1):

        value = ws.cell(
            row=1,
            column=col
        ).value

        if value:
            headers[str(value).strip()] = col

    return headers


def validate_file_name_header(ws):

    headers = get_worksheet_headers(ws)

    if "File Name" not in headers:

        raise Exception(
            "Missing header : File Name"
        )

    return headers


def select_text_column(headers):

    options = {

        name: column

        for name, column in headers.items()

        if name != "File Name"

    }

    if not options:

        raise Exception(
            "No text columns found besides File Name."
        )

    indexed = {}

    print("\nText Columns")

    for index, name in enumerate(sorted(options.keys()), 1):

        print(f"{index}. {name}")
        indexed[str(index)] = name

    choice = ask_choice(
        "\nSelect text column : ",
        indexed.keys()
    )

    return indexed[choice]


def get_tts_column_name(job):

    if job["mode"] == "A":
        return job["text_column"]

    return "Translation"


# ============================================================
#                 WHISPER / SARVAM
# ============================================================

def load_whisper(model_name):

    print(f"\nLoading Whisper ({model_name})...")

    model = WhisperModel(
        model_name,
        device="cpu",
        compute_type="int8"
    )

    print("Whisper Ready.")

    return model


def create_sarvam_client(api_key):

    return SarvamAI(
        api_subscription_key=api_key
    )

    
def convert_audio(input_file, output_format):

    if output_format == "wav":
        return input_file

    output_file = os.path.splitext(input_file)[0] + f".{output_format}"

    subprocess.run(
        [
            "ffmpeg",
            "-y",
            "-i",
            input_file,
            output_file
        ],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        check=True
    )

    os.remove(input_file)

    return output_file


# ============================================================
#                  MASTER EXCEL
# ============================================================

def create_master_excel(project):

    path = master_excel(project)

    if os.path.exists(path):
        return

    wb = Workbook()

    ws = wb.active

    ws.append(MASTER_HEADERS)

    wb.save(path)

    wb.close()


def load_master(project):

    wb = load_workbook(
        master_excel(project)
    )

    return wb, wb.active


# ============================================================
#                  AUDIO DISCOVERY
# ============================================================

def get_audio_files(project):

    return sorted([

        os.path.join(project, file)

        for file in os.listdir(project)

        if file.lower().endswith(
            AUDIO_EXTENSIONS
        )

    ])


# ============================================================
#             AUDIO -> MASTER TRANSCRIPT
# ============================================================

def transcribe_audio(
    project,
    stt_engine,
    whisper=None,
    client=None
):

    wb, ws = load_master(project)

    headers = validate_headers(
        ws,
        MASTER_HEADERS
    )

    file_col = headers["File Name"]
    text_col = headers["Source Transcript"]

    completed = {

        ws.cell(
            row=row,
            column=file_col
        ).value

        for row in range(
            2,
            ws.max_row + 1
        )

    }

    audio_files = get_audio_files(project)

    total = len(audio_files)

    skipped = 0

    for index, audio in enumerate(audio_files, 1):

        filename = os.path.basename(audio)

        if filename in completed:
            skipped += 1
            continue

        
        print(f"[{index}/{total}] {filename}")

        if stt_engine == "whisper":

            segments, info = whisper.transcribe(
                audio,
                beam_size=5,
                vad_filter=True
            )

            transcript = " ".join(
                s.text.strip()
                for s in segments
            )

        else:

            transcript = saaras_transcribe(
                audio,
                client
            )

        ws.append([
            filename,
            transcript
        ])

        wb.save(master_excel(project))

    wb.close()

    print(f"Skipped : {skipped}")


# ============================================================
#          TRANSCRIPT -> MASTER TRANSCRIPT
# ============================================================

def import_transcript_excel(project, transcript_excel):
    
    source = load_workbook(transcript_excel)

    source_ws = source.active

    source_headers = validate_file_name_header(source_ws)

    if "Source Transcript" not in source_headers:

        raise Exception(
            "Missing header : Source Transcript"
        )

    master_wb, master_ws = load_master(
        project
    )

    master_headers = validate_headers(

        master_ws,

        MASTER_HEADERS

    )

    existing = {

        master_ws.cell(

            row=row,

            column=master_headers[
                "File Name"
            ]

        ).value

        for row in range(
            2,
            master_ws.max_row + 1
        )

    }

    for row in range(
        2,
        source_ws.max_row + 1
    ):

        filename = source_ws.cell(

            row=row,

            column=source_headers[
                "File Name"
            ]

        ).value

        transcript = source_ws.cell(

            row=row,

            column=source_headers[
                "Source Transcript"
            ]

        ).value

        if filename in existing:
            continue

        master_ws.append([

            filename,

            transcript

        ])

    master_wb.save(
        master_excel(project)
    )

    source.close()
    master_wb.close()

# ============================================================
#              LANGUAGE EXCEL CREATION
# ============================================================

def create_language_excel(job):

    if os.path.exists(job["excel"]):
        return

    wb = Workbook()

    ws = wb.active

    ws.append([
        "File Name",
        "Source Transcript",
        "Translation"
    ])

    wb.save(job["excel"])
    wb.close()


def populate_language_excel(transcript_excel, job, input_type="audio"):

    create_language_excel(job)
    
    master_wb = load_workbook(transcript_excel)
    master_ws = master_wb.active

    lang_wb = load_workbook(job["excel"])
    lang_ws = lang_wb.active

    if input_type == "transcript":

        source_headers = validate_file_name_header(master_ws)

        if job["text_column"] not in source_headers:

            raise Exception(
                f"Missing header : {job['text_column']}"
            )

        file_col = source_headers["File Name"]
        source_text_col = source_headers[job["text_column"]]

    else:

        master_headers = validate_headers(
            master_ws,
            MASTER_HEADERS
        )

        file_col = master_headers["File Name"]
        source_text_col = master_headers["Source Transcript"]

    lang_headers = validate_headers(
        lang_ws,
        LANGUAGE_HEADERS
    )

    existing = {

        lang_ws.cell(
            row=row,
            column=lang_headers["File Name"]
        ).value

        for row in range(
            2,
            lang_ws.max_row + 1
        )

    }

    changed = False

    for row in range(
        2,
        master_ws.max_row + 1
    ):

        filename = master_ws.cell(
            row=row,
            column=file_col
        ).value

        transcript = master_ws.cell(
            row=row,
            column=source_text_col
        ).value

        if filename in existing:
            continue

        lang_ws.append([
            filename,
            transcript,
            ""
        ])

        changed = True

    if changed:
        lang_wb.save(job["excel"])

    master_wb.close()
    lang_wb.close()


# ============================================================
#                    TRANSLATION
# ============================================================

def translate_language(client, job):

    if job["mode"] == "A":
        return

    wb = load_workbook(job["excel"])
    ws = wb.active

    headers = validate_headers(
        ws,
        LANGUAGE_HEADERS
    )

    file_col = headers["File Name"]
    text_col = headers["Source Transcript"]
    trans_col = headers["Translation"]

    total = ws.max_row - 1
    skipped = 0
    completed = 0

    for row in range(2, ws.max_row + 1):

        filename = ws.cell(
            row=row,
            column=file_col
        ).value

        source_text = ws.cell(
            row=row,
            column=text_col
        ).value

        translated = ws.cell(
            row=row,
            column=trans_col
        ).value

        if translated:

            skipped += 1
            continue
        if not source_text or not str(source_text).strip():
            skipped += 1
            continue

        print(
            f"[{job['name']}] "
            f"{completed+1}/{total} : "
            f"{filename}"
        )

        try:

            if job["transcript_language"] == job["code"]:

                ws.cell(
                    row=row,
                    column=trans_col
                ).value = source_text

                wb.save(job["excel"])

                completed += 1
                continue

            response = client.text.translate(
                input=source_text,
                source_language_code=job["transcript_language"],
                target_language_code=job["code"],
                model="mayura:v1",
                mode=job["translation_mode"]
            )

            ws.cell(
                row=row,
                column=trans_col
            ).value = response.translated_text

            wb.save(job["excel"])

            completed += 1

        except Exception as e:

            print(e)

    wb.close()

    print(
        f"{job['name']} "
        f"Completed:{completed} "
        f"Skipped:{skipped}"
    )


# ============================================================
#                  AUDIO PREVIEW / REVIEW
# ============================================================

def open_preview_folder(path):

    if sys.platform == "win32":
        os.startfile(path)
        return

    if sys.platform == "darwin":
        subprocess.run(["open", path], check=False)
        return

    subprocess.run(["xdg-open", path], check=False)


def parse_regenerate_indices(raw, total):

    if not raw:
        return []

    indices = []

    for part in raw.replace(" ", "").split(","):

        if not part.isdigit():
            raise ValueError(f"Invalid entry : {part}")

        index = int(part)

        if index < 1 or index > total:
            raise ValueError(f"Out of range : {index}")

        indices.append(index)

    return sorted(set(indices))


def get_preview_files(job, audio_format):

    folder = job["preview_folder"]

    if not os.path.isdir(folder):
        return []

    extension = f".{audio_format}"

    return sorted([

        name for name in os.listdir(folder)

        if name.lower().endswith(extension)

    ])


def find_tts_text_for_preview(job, preview_name):

    stem = os.path.splitext(preview_name)[0]
    prefix = f"{job['prefix']}_"

    if not stem.startswith(prefix):
        return None, None

    base = stem[len(prefix):]

    wb = load_workbook(job["excel"])
    ws = wb.active

    headers = validate_headers(
        ws,
        LANGUAGE_HEADERS
    )

    file_col = headers["File Name"]
    tts_col = headers[get_tts_column_name(job)]

    for row in range(2, ws.max_row + 1):

        filename = ws.cell(
            row=row,
            column=file_col
        ).value

        if not filename:
            continue

        if os.path.splitext(str(filename))[0] == base:

            tts_text = ws.cell(
                row=row,
                column=tts_col
            ).value

            wb.close()
            return base, tts_text

    wb.close()
    return base, None


def synthesize_preview_audio(client, job, base, tts_text, audio_format):

    temp_output = os.path.join(
        job["preview_folder"],
        f"{job['prefix']}_{base}.wav"
    )

    audio = client.text_to_speech.convert(
        text=tts_text,
        target_language_code=job["code"],
        model="bulbul:v3",
        speaker=job["speaker"],
        pace=job["pace"]
    )

    save(audio, temp_output)

    return convert_audio(
        temp_output,
        audio_format
    )


def finalize_preview_file(job, preview_name):

    source = os.path.join(
        job["preview_folder"],
        preview_name
    )

    destination = os.path.join(
        job["audio_folder"],
        preview_name
    )

    shutil.move(source, destination)


def review_generated_audio(client, job, audio_format):

    if job["mode"] == "T":
        return

    while True:

        preview_files = get_preview_files(job, audio_format)

        if not preview_files:
            return

        print(f"\n========== {job['name']} Audio Review ==========")
        print("Preview files are ready. Click any file in the folder to play it.\n")

        for index, name in enumerate(preview_files, 1):

            path = os.path.join(
                job["preview_folder"],
                name
            )

            print(f"{index}. {name}")
            print(f"   {path}")

        print(f"\nOpening preview folder...")
        open_preview_folder(job["preview_folder"])
        print(
            "Double-click files in File Explorer to listen. "
            "Come back here when you're done."
        )

        input("\nPress Enter when you're finished listening...")

        while True:

            raw = input(
                "\nEnter comma-separated numbers to REGENERATE "
                "(empty = save all remaining) : "
            ).strip()

            try:
                regenerate_indices = parse_regenerate_indices(
                    raw,
                    len(preview_files)
                )
                break
            except ValueError as error:
                print(error)

        if not regenerate_indices:

            for name in preview_files:
                finalize_preview_file(job, name)
                print(f"Saved : {name}")

            return

        regenerate_names = {

            preview_files[index - 1]

            for index in regenerate_indices

        }

        for name in preview_files:

            if name in regenerate_names:
                continue

            finalize_preview_file(job, name)
            print(f"Saved : {name}")

        for name in sorted(regenerate_names):

            base, tts_text = find_tts_text_for_preview(
                job,
                name
            )

            preview_path = os.path.join(
                job["preview_folder"],
                name
            )

            if os.path.exists(preview_path):
                os.remove(preview_path)

            if not tts_text:

                print(f"Cannot regenerate (missing text) : {name}")
                continue

            print(f"Regenerating : {name}")

            try:

                synthesize_preview_audio(
                    client,
                    job,
                    base,
                    tts_text,
                    audio_format
                )

            except Exception as error:
                print(error)


# ============================================================
#                       TTS
# ============================================================

def generate_audio(client, job, audio_format):

    if job["mode"] == "T":
        return

    wb = load_workbook(job["excel"])
    ws = wb.active

    headers = validate_headers(
        ws,
        LANGUAGE_HEADERS
    )

    file_col = headers["File Name"]
    tts_col = headers[get_tts_column_name(job)]

    total = ws.max_row - 1
    skipped = 0
    completed = 0

    for row in range(2, ws.max_row + 1):

        filename = ws.cell(
            row=row,
            column=file_col
        ).value

        tts_text = ws.cell(
            row=row,
            column=tts_col
        ).value

        if not tts_text:

            print(
                f"[{job['name']}] "
                f"Missing text : "
                f"{filename}"
            )

            continue

        base = os.path.splitext(filename)[0]

        temp_output = os.path.join(
            job["preview_folder"],
            f"{job['prefix']}_{base}.wav"
        )

        preview_output = os.path.join(
            job["preview_folder"],
            f"{job['prefix']}_{base}.{audio_format}"
        )

        final_output = os.path.join(
            job["audio_folder"],
            f"{job['prefix']}_{base}.{audio_format}"
        )

        if os.path.exists(final_output):
            skipped += 1
            continue

        if os.path.exists(preview_output):
            skipped += 1
            continue

        print(
            f"[{job['name']}] "
            f"{completed+1}/{total} : "
            f"{filename}"
        )

        try:

            audio = client.text_to_speech.convert(
                text=tts_text,
                target_language_code=job["code"],
                model="bulbul:v3",
                speaker=job["speaker"],
                pace=job["pace"]
            )

            save(audio, temp_output)

            convert_audio(
                temp_output,
                audio_format
            )

            completed += 1

        except Exception as e:
            print(e)
            
def process_language(api_key, project, transcript_excel, job, audio_format, input_type="audio"):

    try:

        client = create_sarvam_client(api_key)

        populate_language_excel(
              transcript_excel,
                job,
                input_type
               )

        translate_language(
            client,
            job
        )

        generate_audio(
            client,
            job,
            audio_format
        )

        review_generated_audio(
            client,
            job,
            audio_format
        )

        print(
            f"{job['name']} Finished."
        )

    except Exception as e:

        print(
            f"{job['name']} Failed : {e}"
        )


# ============================================================
#                PIPELINE RUNNER
# ============================================================

def run_pipeline(

    api_key,

    project,

    jobs,
    
    transcript_excel,
    
    audio_format,

    input_type="audio"
    
  ):

    for job in jobs:

        process_language(

            api_key,

            project,
                
            transcript_excel,

            job,
                
            audio_format,

            input_type

        )


# ============================================================
#                        MAIN
# ============================================================

def main():

    input_type = get_input_type()
    
    transcript_language = None

    transcript_excel = None

    if input_type == "audio":
        stt_engine = None
    if input_type == "audio":
        stt_engine = get_stt_engine()
        project = get_project_folder()
    else:
        transcript_excel = get_transcript_excel()
        project = os.path.dirname(transcript_excel)

    project_mode = get_project_mode(project)

    if input_type == "audio":
      create_master_excel(project)

    if input_type == "audio":
        validate_audio_project(project)
        

    whisper = None
    stt_client = None

    if input_type == "audio":

        if stt_engine == "whisper":

            whisper_model = get_whisper_model()
            whisper = load_whisper(whisper_model)

        else:

            api_key = get_api_key()
            stt_client = create_sarvam_client(api_key)

    

            if input_type == "audio":

                transcribe_audio(
                            project,
                            stt_engine,
                            whisper,
                            stt_client
                        )

    
            
    if input_type == "audio":
        transcript_excel = master_excel(project)

    api_key = input(
        "\nSarvam API Key (Press Enter to skip): "
    ).strip()

    if not api_key:

        print("\nSkipping Translation / TTS.")
        print(" Pipeline Completed Successfully ")

        return

    language_ids = get_target_languages()

    jobs = configure_languages(
        language_ids,
        transcript_excel,
        input_type
    )
    
    needs_translation = any(
    job["mode"] in ("T", "TA")
    for job in jobs
)

    if needs_translation:
        transcript_language = get_transcript_language()
        translation_mode = get_translation_mode()
    else:
        transcript_language = None
        translation_mode = None

    for job in jobs:
        job["translation_mode"] = translation_mode
        job["transcript_language"] = transcript_language

    audio_format = get_audio_format()

    create_output_structure(
        project,
        jobs
    )

    run_pipeline(
        api_key,
        project,
        jobs,
        transcript_excel,
        audio_format,
        input_type
    )

    
    print(" Pipeline Completed Successfully ")
    



if __name__ == "__main__":

    main()

# =================== END OF FILE ===================